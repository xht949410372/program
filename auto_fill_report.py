import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy
import random
import os
import configparser
import tkinter as tk
from tkinter import filedialog

def read_config():
    """读取配置文件"""
    config = configparser.ConfigParser()
    config_file = 'config.ini'
    
    # 如果配置文件不存在，创建默认配置
    if not os.path.exists(config_file):
        config['PackageColumns'] = {
            'names': '包号,件号,Roll No,Package No,卷号,编号',
            'quantity_names': '数量,件数,码数'
        }
        config['FormatOptions'] = {
            'add_prefixes': 'false',
            'include_actual_length': 'true',
            'min_score': '5',
            'max_score': '14',
            'score_items_count': '28',
            'empty_lines_between_total_and_other': '1'
        }
        with open(config_file, 'w', encoding='utf-8') as f:
            config.write(f)
    
    # 读取配置，指定utf-8编码
    config.read(config_file, encoding='utf-8')
    package_names = config['PackageColumns']['names'].split(',')
    quantity_names = config['PackageColumns'].get('quantity_names', '数量,件数,码数').split(',')
    add_prefixes = config.getboolean('FormatOptions', 'add_prefixes', fallback=False)
    include_actual_length = config.getboolean('FormatOptions', 'include_actual_length', fallback=True)
    min_score = config.getint('FormatOptions', 'min_score', fallback=5)
    max_score = config.getint('FormatOptions', 'max_score', fallback=14)
    score_items_count = config.getint('FormatOptions', 'score_items_count', fallback=28)
    empty_lines = config.getint('FormatOptions', 'empty_lines_between_total_and_other', fallback=1)
    return package_names, quantity_names, add_prefixes, include_actual_length, min_score, max_score, score_items_count, empty_lines

def read_adw70_data(file_path):
    """读取Excel文件中的数据，按sheet分组，支持.xls和.xlsx格式，智能识别列标题"""
    import os
    sheet_data = {}
    
    # 读取配置文件中的包号标识符
    package_names, quantity_names, _, _, _, _, _, _ = read_config()
    
    # 根据文件扩展名选择合适的库
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.xlsx':
        # 使用openpyxl读取.xlsx文件
        wb = openpyxl.load_workbook(file_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = []
            has_batch_no = False
            
            # 尝试识别表头行
            header_row = None
            header_row_idx = -1
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
                if any(cell and any(pkg_name in str(cell) for pkg_name in package_names) for cell in row):
                    header_row = row
                    header_row_idx = row_idx + 1  # 转换为Excel行号（从1开始）
                    break
            
            if header_row:
                # 识别包号/件号列、数量列和缸号列
                package_columns = []
                quantity_columns = []
                batch_columns = []
                
                for i, cell in enumerate(header_row):
                    if cell and any(pkg_name in str(cell) for pkg_name in package_names):
                        package_columns.append(i)
                    elif cell and any(qty_name in str(cell) for qty_name in quantity_names):
                        quantity_columns.append(i)
                    elif cell and ('缸号' in str(cell) or 'lot' in str(cell).lower() or 'batch' in str(cell).lower()):
                        batch_columns.append(i)
                
                print(f"在{sheet_name}中识别到包号/件号列: {package_columns}")
                print(f"在{sheet_name}中识别到数量列: {quantity_columns}")
                print(f"在{sheet_name}中识别到缸号列: {batch_columns}")
                
                has_batch_no = len(batch_columns) > 0
                
                # 读取数据，先竖着读每一列的包号/件号
                start_row = header_row_idx + 1
                max_row = ws.max_row
                
                # 按列处理
                for pkg_col in package_columns:
                    # 查找对应的数量列
                    quantity_col = None
                    for q_col in quantity_columns:
                        if q_col > pkg_col:
                            quantity_col = q_col
                            break
                    
                    # 查找对应的缸号列
                    batch_col = None
                    for b_col in batch_columns:
                        if b_col > pkg_col:
                            batch_col = b_col
                            break
                    
                    if quantity_col:
                        # 处理该列的所有行
                        for row in ws.iter_rows(min_row=start_row, max_row=max_row, values_only=True):
                            if pkg_col < len(row) and row[pkg_col] is not None:
                                try:
                                    # 处理件号格式，移除#号
                                    pkg_value = str(row[pkg_col])
                                    pkg_value = pkg_value.replace('#', '').strip()
                                    package_num = int(pkg_value)
                                    
                                    # 处理数量值
                                    if quantity_col < len(row):
                                        qty_value = row[quantity_col]
                                        if qty_value not in (None, ''):
                                            quantity = float(qty_value)
                                            
                                            # 处理缸号值
                                            batch_no = None
                                            if batch_col and batch_col < len(row):
                                                batch_value = row[batch_col]
                                                if batch_value not in (None, ''):
                                                    batch_no = str(batch_value).strip()
                                            
                                            data.append((package_num, quantity, batch_no))
                                except Exception as e:
                                    # 跳过无法处理的行
                                    pass
            else:
                # 如果没有识别到表头，使用默认列
                print(f"在{sheet_name}中未识别到表头，使用默认列")
                for row in ws.iter_rows(min_row=4, values_only=True):
                    if row[0] is not None:
                        try:
                            # 处理件号格式，移除#号
                            pkg_value = str(row[0])
                            pkg_value = pkg_value.replace('#', '').strip()
                            package_num = int(pkg_value)
                            quantity = float(row[2])
                            # 假设默认没有缸号
                            data.append((package_num, quantity, None))
                        except Exception as e:
                            # 跳过无法处理的行
                            pass
                    # 检查其他包号列
                    if row[3] is not None:
                        try:
                            # 处理件号格式，移除#号
                            pkg_value = str(row[3])
                            pkg_value = pkg_value.replace('#', '').strip()
                            package_num = int(pkg_value)
                            quantity = float(row[5])
                            # 假设默认没有缸号
                            data.append((package_num, quantity, None))
                        except Exception as e:
                            # 跳过无法处理的行
                            pass
                    if row[6] is not None:
                        try:
                            # 处理件号格式，移除#号
                            pkg_value = str(row[6])
                            pkg_value = pkg_value.replace('#', '').strip()
                            package_num = int(pkg_value)
                            quantity = float(row[8])
                            # 假设默认没有缸号
                            data.append((package_num, quantity, None))
                        except Exception as e:
                            # 跳过无法处理的行
                            pass
            
            # 存储该sheet的数据和是否有缸号
            sheet_data[sheet_name] = (data, has_batch_no)
        
        wb.close()
        
    elif ext == '.xls':
        # 使用xlrd读取.xls文件
        wb = xlrd.open_workbook(file_path)
        
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            data = []
            has_batch_no = False
            
            # 尝试识别表头行
            header_row = None
            header_row_idx = -1
            for row_idx in range(min(10, ws.nrows)):
                row = ws.row_values(row_idx)
                if any(cell and any(pkg_name in str(cell) for pkg_name in package_names) for cell in row):
                    header_row = row
                    header_row_idx = row_idx
                    break
            
            if header_row:
                # 识别包号/件号列、数量列和缸号列
                package_columns = []
                quantity_columns = []
                batch_columns = []
                
                for i, cell in enumerate(header_row):
                    if cell and any(pkg_name in str(cell) for pkg_name in package_names):
                        package_columns.append(i)
                    elif cell and any(qty_name in str(cell) for qty_name in quantity_names):
                        quantity_columns.append(i)
                    elif cell and ('缸号' in str(cell) or 'lot' in str(cell).lower() or 'batch' in str(cell).lower()):
                        batch_columns.append(i)
                
                print(f"在{sheet_name}中识别到包号/件号列: {package_columns}")
                print(f"在{sheet_name}中识别到数量列: {quantity_columns}")
                print(f"在{sheet_name}中识别到缸号列: {batch_columns}")
                
                has_batch_no = len(batch_columns) > 0
                
                # 读取数据，先竖着读每一列的包号/件号
                start_row = header_row_idx + 1
                max_row = ws.nrows
                
                # 按列处理
                for pkg_col in package_columns:
                    # 查找对应的数量列
                    quantity_col = None
                    for q_col in quantity_columns:
                        if q_col > pkg_col:
                            quantity_col = q_col
                            break
                    
                    # 查找对应的缸号列
                    batch_col = None
                    for b_col in batch_columns:
                        if b_col > pkg_col:
                            batch_col = b_col
                            break
                    
                    if quantity_col:
                        # 处理该列的所有行
                        for row_idx in range(start_row, max_row):
                            row = ws.row_values(row_idx)
                            if pkg_col < len(row) and row[pkg_col] not in (None, ''):
                                try:
                                    # 处理件号格式，移除#号
                                    pkg_value = str(row[pkg_col])
                                    pkg_value = pkg_value.replace('#', '').strip()
                                    package_num = int(pkg_value)
                                    
                                    # 处理数量值
                                    if quantity_col < len(row):
                                        qty_value = row[quantity_col]
                                        if qty_value not in (None, ''):
                                            quantity = float(qty_value)
                                            
                                            # 处理缸号值
                                            batch_no = None
                                            if batch_col and batch_col < len(row):
                                                batch_value = row[batch_col]
                                                if batch_value not in (None, ''):
                                                    batch_no = str(batch_value).strip()
                                            
                                            data.append((package_num, quantity, batch_no))
                                except Exception as e:
                                    # 跳过无法处理的行
                                    pass
            else:
                # 如果没有识别到表头，使用默认列
                print(f"在{sheet_name}中未识别到表头，使用默认列")
                for row_idx in range(3, ws.nrows):  # 3是第4行的索引
                    row = ws.row_values(row_idx)
                    if row[0] not in (None, ''):
                        try:
                            # 处理件号格式，移除#号
                            pkg_value = str(row[0])
                            pkg_value = pkg_value.replace('#', '').strip()
                            package_num = int(pkg_value)
                            quantity = float(row[2])
                            # 假设默认没有缸号
                            data.append((package_num, quantity, None))
                        except Exception as e:
                            # 跳过无法处理的行
                            pass
                    # 检查其他包号列
                    if row[3] not in (None, ''):
                        try:
                            # 处理件号格式，移除#号
                            pkg_value = str(row[3])
                            pkg_value = pkg_value.replace('#', '').strip()
                            package_num = int(pkg_value)
                            quantity = float(row[5])
                            # 假设默认没有缸号
                            data.append((package_num, quantity, None))
                        except Exception as e:
                            # 跳过无法处理的行
                            pass
                    if row[6] not in (None, ''):
                        try:
                            # 处理件号格式，移除#号
                            pkg_value = str(row[6])
                            pkg_value = pkg_value.replace('#', '').strip()
                            package_num = int(pkg_value)
                            quantity = float(row[8])
                            # 假设默认没有缸号
                            data.append((package_num, quantity, None))
                        except Exception as e:
                            # 跳过无法处理的行
                            pass
            
            # 存储该sheet的数据和是否有缸号
            sheet_data[sheet_name] = (data, has_batch_no)
        
        # 检查xlrd的Book对象是否有close方法
        if hasattr(wb, 'close'):
            wb.close()
    else:
        print(f"不支持的文件格式: {ext}")
    
    return sheet_data

def generate_random_score():
    """生成随机分数，5-14分"""
    return random.randint(5, 14)

def calculate_total_score(scores):
    """计算总分数"""
    return sum(scores)

def fill_report(source_file):
    """根据源文件数据生成只包含吊码长、实际码长和计分的Excel文件，每个sheet生成一个文件"""
    try:
        print("开始处理...")
        # 读取配置文件
        _, _, add_prefixes, include_actual_length, min_score, max_score, score_items_count, empty_lines = read_config()
        # 读取ADW70数据，按sheet分组
        sheet_data = read_adw70_data(source_file)
        print(f"读取到 {len(sheet_data)} 个sheet")
        
        # 定义计分项目
        score_items = [
            '横档',
            '色点',
            '断经',
            '断纬',
            '纱节',
            '破洞',
            '色档',
            '色花',
            '筘路',
            '露白',
            '污渍',
            '油污',
            '搭色',
            '刀口伤',
            '擦伤',
            '接匹',
            '浆斑',
            '粗纱',
            '钩丝',
            '死皱',
            '隐档',
            '手感',
            '纬斜',
            '边中色差',
            '头尾色差',
            '其他'
        ]
        
        # 根据配置调整计分项目数量
        if score_items_count < len(score_items):
            # 如果配置的数量少于默认数量，去掉后面的项目
            score_items = score_items[:score_items_count]
        elif score_items_count > len(score_items):
            # 如果配置的数量多于默认数量，添加更多的'其他'项目
            additional_items = score_items_count - len(score_items)
            for i in range(additional_items):
                score_items.append('其他')
        
        # 为每个sheet生成一个Excel文件
        for sheet_name, (adw70_data, has_batch_no) in sheet_data.items():
            print(f"\n处理sheet: {sheet_name}")
            print(f"读取到 {len(adw70_data)} 条数据")
            print(f"是否有缸号: {has_batch_no}")
            
            # 检查数据格式
            if adw70_data:
                print(f"第一条数据: {adw70_data[0]}")
                if len(adw70_data[0]) >= 2:
                    print(f"数据类型: package_num={type(adw70_data[0][0])}, quantity={type(adw70_data[0][1])}")
                if len(adw70_data[0]) >= 3:
                    print(f"缸号: {adw70_data[0][2]}")
            
            # 创建新的工作簿
            print("创建新工作簿...")
            new_workbook = xlwt.Workbook()
            new_worksheet = new_workbook.add_sheet('数据')
            print("新工作簿创建成功")
            
            # 创建居中样式
            center_style = xlwt.XFStyle()
            alignment = xlwt.Alignment()
            alignment.horz = xlwt.Alignment.HORZ_CENTER  # 水平居中
            alignment.vert = xlwt.Alignment.VERT_CENTER  # 垂直居中
            center_style.alignment = alignment
            
            # 填写卷号和吊码长、实际码长
            roll_data = []
            batch_data = []  # 存储缸号数据
            
            print("填写卷号数据...")
            for i, item in enumerate(adw70_data):
                if len(item) >= 3:
                    package_num, quantity, batch_no = item
                else:
                    package_num, quantity = item
                    batch_no = None
                
                # 计算实际码长（吊码长基础上随机增加0.1-0.5）
                actual_length = quantity + round(random.uniform(0.1, 0.5), 1)
                
                # 存储卷号数据，直接使用包号作为卷号
                roll_data.append((package_num, quantity, actual_length))
                batch_data.append((package_num, quantity, batch_no))
                print(f"第 {package_num} 卷: 吊码长={quantity}, 实际码长={actual_length}, 缸号={batch_no}")
            
            # 计算总卷数
            total_rolls = len(roll_data)
            print(f"总卷数: {total_rolls}")
            
            # 每5卷一组
            groups = (total_rolls + 4) // 5
            print(f"分组数: {groups}")
            
            # 填写数据和计分
            print("填写计分数据...")
            
            # 为每个卷单独计算分数
            roll_scores = []
            
            # 每5卷一组处理
            for group in range(groups):
                # 计算起始卷号和结束卷号
                start_roll = group * 5
                end_roll = min((group + 1) * 5, total_rolls)
                
                # 根据是否包含实际码长计算行偏移
                if include_actual_length:
                    row_offset = group * (len(score_items) + 9)  # 包含实际码长时，每组之间空两行
                else:
                    row_offset = group * (len(score_items) + 8)  # 不包含实际码长时，每组之间空两行
                
                # 填写标题
                new_worksheet.write(row_offset, 0, '卷号')
                new_worksheet.write(row_offset + 1, 0, '吊码长')
                if include_actual_length:
                    new_worksheet.write(row_offset + 2, 0, '实际码长')
                    new_worksheet.write(row_offset + 3, 0, '门幅')
                    new_worksheet.write(row_offset + 4, 0, '计分')
                else:
                    new_worksheet.write(row_offset + 2, 0, '门幅')
                    new_worksheet.write(row_offset + 3, 0, '计分')
                
                # 根据是否包含实际码长确定计分项目名称的起始行
                if include_actual_length:
                    score_start_row = row_offset + 5
                else:
                    score_start_row = row_offset + 4
                
                # 填写计分列标题（1-4）
                for i in range(end_roll - start_roll):
                    if include_actual_length:
                        new_worksheet.write(row_offset + 4, 1 + i * 4, 1)
                        new_worksheet.write(row_offset + 4, 2 + i * 4, 2)
                        new_worksheet.write(row_offset + 4, 3 + i * 4, 3)
                        new_worksheet.write(row_offset + 4, 4 + i * 4, 4)
                    else:
                        new_worksheet.write(row_offset + 3, 1 + i * 4, 1)
                        new_worksheet.write(row_offset + 3, 2 + i * 4, 2)
                        new_worksheet.write(row_offset + 3, 3 + i * 4, 3)
                        new_worksheet.write(row_offset + 3, 4 + i * 4, 4)
                
                # 先填写计分项目名称
                for row_idx, item in enumerate(score_items, start=score_start_row):
                    new_worksheet.write(row_idx, 0, item)
                
                # 填写卷号数据
                for i in range(start_roll, end_roll):
                    package_num, quantity, actual_length = roll_data[i]
                    group_idx = i - start_roll
                    
                    # 生成随机门幅（142-142.5），保留一位小数
                    width = round(random.uniform(142, 142.5), 1)
                    
                    # 合并单元格：卷号、吊码长、实际码长、门幅对应的列
                    # 合并卷号单元格（4列合并成一个）
                    new_worksheet.write_merge(row_offset, row_offset, 1 + group_idx * 4, 4 + group_idx * 4, f"{package_num}#", style=center_style)
                    
                    # 根据配置决定是否添加前缀
                    if add_prefixes:
                        # 合并吊码长单元格（4列合并成一个）
                        new_worksheet.write_merge(row_offset + 1, row_offset + 1, 1 + group_idx * 4, 4 + group_idx * 4, f"WIDTH: {quantity}", style=center_style)
                        # 根据配置决定是否填写实际码长
                        if include_actual_length:
                            # 合并实际码长单元格（4列合并成一个）
                            new_worksheet.write_merge(row_offset + 2, row_offset + 2, 1 + group_idx * 4, 4 + group_idx * 4, f"WIDTH: {actual_length}", style=center_style)
                            # 合并门幅单元格（4列合并成一个）
                            new_worksheet.write_merge(row_offset + 3, row_offset + 3, 1 + group_idx * 4, 4 + group_idx * 4, f"YARDS: {width}", style=center_style)
                        else:
                            # 合并门幅单元格（4列合并成一个）
                            new_worksheet.write_merge(row_offset + 2, row_offset + 2, 1 + group_idx * 4, 4 + group_idx * 4, f"YARDS: {width}", style=center_style)
                    else:
                        # 合并吊码长单元格（4列合并成一个）
                        new_worksheet.write_merge(row_offset + 1, row_offset + 1, 1 + group_idx * 4, 4 + group_idx * 4, quantity, style=center_style)
                        # 根据配置决定是否填写实际码长
                        if include_actual_length:
                            # 合并实际码长单元格（4列合并成一个）
                            new_worksheet.write_merge(row_offset + 2, row_offset + 2, 1 + group_idx * 4, 4 + group_idx * 4, actual_length, style=center_style)
                            # 合并门幅单元格（4列合并成一个）
                            new_worksheet.write_merge(row_offset + 3, row_offset + 3, 1 + group_idx * 4, 4 + group_idx * 4, width, style=center_style)
                        else:
                            # 合并门幅单元格（4列合并成一个）
                            new_worksheet.write_merge(row_offset + 2, row_offset + 2, 1 + group_idx * 4, 4 + group_idx * 4, width, style=center_style)
                
                # 处理该组的卷
                for i in range(start_roll, end_roll):
                    package_num, _, _ = roll_data[i]
                    group_idx = i - start_roll
                    
                    # 确定每卷的目标分数（用户指定范围）
                    target_score = random.randint(min_score, max_score)
                    remaining_score = target_score
                    
                    # 存储已填写的问题索引和位置
                    filled_positions = set()
                    
                    # 随机填写分数，直到没有剩余分数
                    while remaining_score > 0:
                        # 随机选择一个问题和列位置
                        available_questions = list(range(len(score_items)))
                        available_columns = list(range(4))  # 4列
                        
                        # 生成所有可能的位置组合
                        possible_positions = [(q, c) for q in available_questions for c in available_columns if (q, c) not in filled_positions]
                        
                        if not possible_positions:
                            break
                        
                        # 随机选择一个位置
                        item_idx, col_offset = random.choice(possible_positions)
                        filled_positions.add((item_idx, col_offset))
                        row_idx = score_start_row + item_idx
                        
                        # 尝试填写分数，在1-3之间随机选择
                        score = random.randint(1, 3)
                        # 计算该位置的分数值：score * 列分数（列标题）
                        col_score = col_offset + 1  # 列标题是1-4
                        score_value = score * col_score
                        
                        if score_value > remaining_score:
                            # 如果超过剩余分数，尝试更小的分数
                            for s in range(score - 1, 0, -1):
                                score_value = s * col_score
                                if score_value <= remaining_score:
                                    score = s
                                    break
                            else:
                                # 如果所有分数都超过，跳过这个位置
                                continue
                        
                        # 填写计分
                        new_worksheet.write(row_idx, 1 + group_idx * 4 + col_offset, score)
                        # 减去已使用的分数
                        remaining_score -= score_value
                    
                    # 确保分数正好用完，如果还有剩余分数，尝试填补
                    if remaining_score > 0:
                        # 重新生成可能的位置
                        possible_positions = [(q, c) for q in range(len(score_items)) for c in range(4) if (q, c) not in filled_positions]
                        
                        for q, c in possible_positions:
                            if remaining_score <= 0:
                                break
                            
                            # 尝试填写1-3之间的随机数
                            max_score = min(3, remaining_score)
                            if max_score < 1:
                                continue
                            
                            score = random.randint(1, max_score)
                            col_score = c + 1  # 列标题是1-4
                            score_value = score * col_score
                            
                            if remaining_score >= score_value:
                                new_worksheet.write(score_start_row + q, 1 + group_idx * 4 + c, score)
                                remaining_score -= score_value
                        
                        # 如果还有剩余分数，调整目标分数
                        if remaining_score > 0:
                            # 重新计算目标分数，确保至少5分
                            target_score = max(5, target_score - remaining_score)
                            remaining_score = 0
                    
                    # 计算实际总分数
                    current_score = target_score - remaining_score
                    
                    # 确保分数在用户指定范围内
                    if current_score < min_score:
                        # 计算需要添加的分数
                        needed_score = min_score - current_score
                        current_score = min_score
                        # 尝试添加分数直到达到5分
                        possible_positions = [(q, c) for q in range(len(score_items)) for c in range(4) if (q, c) not in filled_positions]
                        
                        for q, c in possible_positions:
                            if needed_score <= 0:
                                break
                            # 尝试填写1-3之间的随机数
                            max_score = min(3, needed_score)
                            if max_score < 1:
                                continue
                            
                            score = random.randint(1, max_score)
                            col_score = c + 1  # 列标题是1-4
                            score_value = score * col_score
                            
                            if needed_score >= score_value:
                                new_worksheet.write(score_start_row + q, 1 + group_idx * 4 + c, score)
                                needed_score -= score_value
                    
                    roll_scores.append(current_score)
                    print(f"第 {package_num} 卷扣分: {current_score}")
                
                # 填写该组的总扣分
                print("填写总扣分...")
                # 计算总扣分的行位置，在计分项目最后一行之后添加空行
                total_deduction_row = score_start_row + len(score_items) + empty_lines
                new_worksheet.write(total_deduction_row, 0, '总扣分')
                for i in range(start_roll, end_roll):
                    package_num = roll_data[i][0]
                    group_idx = i - start_roll
                    # 使用正确的索引访问roll_scores
                    score_idx = i
                    if score_idx < len(roll_scores):
                        # 合并总扣分单元格（4列合并成一个）
                        new_worksheet.write_merge(total_deduction_row, total_deduction_row, 1 + group_idx * 4, 4 + group_idx * 4, roll_scores[score_idx], style=center_style)
                    else:
                        # 如果索引超出范围，使用默认值
                        new_worksheet.write_merge(total_deduction_row, total_deduction_row, 1 + group_idx * 4, 4 + group_idx * 4, 5, style=center_style)
                
                # 100码评分在总扣分的下一行
                total_row = total_deduction_row + 1
                
                # 填写100码评分
                new_worksheet.write(total_row, 0, '100码评分')
                for i in range(start_roll, end_roll):
                    package_num, quantity, _ = roll_data[i]
                    group_idx = i - start_roll
                    # 使用正确的索引访问roll_scores
                    score_idx = i
                    if score_idx < len(roll_scores):
                        total_deduction = roll_scores[score_idx]
                        # 计算100码评分：(总扣分*3600)/(吊码长/0.9144)/56
                        yards = quantity / 0.9144  # 转换为码
                        score_100y = (total_deduction * 3600) / yards / 56
                        score_100y = round(score_100y, 1)  # 保留一位小数
                        # 合并100码评分单元格（4列合并成一个）
                        new_worksheet.write_merge(total_row, total_row, 1 + group_idx * 4, 4 + group_idx * 4, score_100y, style=center_style)
                    else:
                        # 如果索引超出范围，使用默认值
                        new_worksheet.write_merge(total_row, total_row, 1 + group_idx * 4, 4 + group_idx * 4, 0, style=center_style)
            print("总扣分填写完成")
            
            # 如果有缸号，添加缸号统计信息
            if has_batch_no:
                print("添加缸号统计信息...")
                # 计算最后一行的位置
                if include_actual_length:
                    last_row = groups * (len(score_items) + 9) - 1
                else:
                    last_row = groups * (len(score_items) + 8) - 1
                
                # 空两行
                batch_start_row = last_row + 3
                
                # 填写缸号统计标题
                new_worksheet.write(batch_start_row, 0, '缸号')
                new_worksheet.write(batch_start_row, 1, '总件数')
                new_worksheet.write(batch_start_row, 2, '总数量')
                
                # 统计缸号数据
                batch_stats = {}
                for package_num, quantity, batch_no in batch_data:
                    if batch_no:
                        if batch_no not in batch_stats:
                            batch_stats[batch_no] = {'count': 0, 'total_quantity': 0}
                        batch_stats[batch_no]['count'] += 1
                        batch_stats[batch_no]['total_quantity'] += quantity
                
                # 填写缸号统计数据
                row_idx = batch_start_row + 1
                for batch_no, stats in batch_stats.items():
                    new_worksheet.write(row_idx, 0, batch_no)
                    new_worksheet.write(row_idx, 1, stats['count'])
                    new_worksheet.write(row_idx, 2, stats['total_quantity'])
                    row_idx += 1
                
                print(f"已添加 {len(batch_stats)} 个缸号的统计信息")
            
            # 另存为新文件，使用sheet名字作为文件名
            output_file = f"{sheet_name}_报告.xls"
            print(f"保存文件到: {output_file}")
            try:
                new_workbook.save(output_file)
                print(f"数据生成完成，已保存到 {output_file}")
            except Exception as e:
                print(f"保存文件时出错: {str(e)}, output_file={output_file}")
                return False
        
        return True
    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def select_file(title, filetypes):
    """打开文件选择对话框"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    file_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    
    # 检查是否选择了.xls文件
    if file_path and file_path.lower().endswith('.xls'):
        # 弹出提示对话框
        from tkinter.messagebox import showinfo
        showinfo("提示", "建议将.xls文件另存为.xlsx格式以获得更好的兼容性和性能。")
        root.destroy()
        return ""  # 返回空字符串，停止执行
    
    root.destroy()
    return file_path

def get_score_range():
    """获取用户输入的分数范围，如果用户取消则使用默认值"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    
    # 创建输入对话框
    from tkinter.simpledialog import askinteger
    
    min_score = askinteger("分数范围", "请输入最低分数:", minvalue=1, maxvalue=20, initialvalue=5)
    if min_score is None:
        root.destroy()
        return 5, 14  # 使用默认值
    
    max_score = askinteger("分数范围", "请输入最高分数:", minvalue=min_score, maxvalue=50, initialvalue=14)
    if max_score is None:
        root.destroy()
        return 5, 14  # 使用默认值
    
    root.destroy()
    return min_score, max_score

def main():
    print("检验报告自动填写工具")
    print("=" * 50)
    print("Author: XHT")
    print("Version: 1.0.0")
    print("Date: 2026-03-31")
    print("=" * 50)
    
    # 选择码单文件
    print("请选择码单文件...")
    source_file = select_file("选择码单文件", [("Excel文件", "*.xlsx;*.xls")])
    if not source_file:
        print("错误: 未选择码单文件")
        return
    print(f"选择的码单文件: {source_file}")
    
    # 检查文件是否存在
    if not os.path.exists(source_file):
        print(f"错误: 源文件 {source_file} 不存在")
        return
    
    # 执行自动填写
    success = fill_report(source_file)
    if success:
        print("操作成功完成")
    else:
        print("操作失败")

if __name__ == "__main__":
    main()